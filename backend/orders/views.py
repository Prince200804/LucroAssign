from rest_framework import generics, status, permissions
from rest_framework.response import Response
from rest_framework.views import APIView
from django.db import transaction
from django.shortcuts import get_object_or_404
from django.conf import settings
from django.http import HttpResponse
from .models import Order, OrderItem
from .serializers import (
    OrderSerializer, CreateOrderSerializer, OrderListSerializer,
    AdminOrderSerializer, OrderStatusUpdateSerializer
)
from cart.models import Cart
from analytics.utils import track_interaction
import stripe
import io
from datetime import datetime


# Stripe configuration
stripe.api_key = getattr(settings, 'STRIPE_SECRET_KEY', 'sk_test_your_secret_key')
STRIPE_PUBLISHABLE_KEY = getattr(settings, 'STRIPE_PUBLISHABLE_KEY', 'pk_test_your_publishable_key')


class CreateStripePaymentIntentView(APIView):
    """Create Stripe Payment Intent for checkout."""
    
    permission_classes = [permissions.AllowAny]
    
    def get_cart(self, request):
        if request.user.is_authenticated:
            try:
                return Cart.objects.get(user=request.user)
            except Cart.DoesNotExist:
                return None
        else:
            session_key = request.session.session_key
            if session_key:
                try:
                    return Cart.objects.get(session_key=session_key)
                except Cart.DoesNotExist:
                    return None
        return None
    
    def post(self, request):
        cart = self.get_cart(request)
        if not cart or cart.items.count() == 0:
            return Response(
                {'error': 'Cart is empty'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check stock availability
        for item in cart.items.all():
            if item.product.stock < item.quantity:
                return Response(
                    {'error': f'Insufficient stock for {item.product.name}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        amount_in_paise = int(float(cart.total) * 100)  # Convert to paise (smallest currency unit)
        
        try:
            # Create Stripe Payment Intent
            intent = stripe.PaymentIntent.create(
                amount=amount_in_paise,
                currency='inr',
                automatic_payment_methods={
                    'enabled': True,
                },
                metadata={
                    'cart_id': str(cart.id) if hasattr(cart, 'id') else '',
                }
            )
            
            return Response({
                'client_secret': intent.client_secret,
                'payment_intent_id': intent.id,
                'publishable_key': STRIPE_PUBLISHABLE_KEY,
                'amount': amount_in_paise,
                'currency': 'inr',
            })
        except stripe.error.StripeError as e:
            return Response(
                {'error': str(e.user_message if hasattr(e, 'user_message') else e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CreateOrderView(APIView):
    """Create order from cart after payment."""
    
    permission_classes = [permissions.AllowAny]
    
    def get_cart(self, request):
        if request.user.is_authenticated:
            try:
                return Cart.objects.get(user=request.user)
            except Cart.DoesNotExist:
                return None
        else:
            session_key = request.session.session_key
            if session_key:
                try:
                    return Cart.objects.get(session_key=session_key)
                except Cart.DoesNotExist:
                    return None
        return None
    
    def verify_stripe_payment(self, payment_intent_id):
        """Verify the Stripe payment status."""
        try:
            intent = stripe.PaymentIntent.retrieve(payment_intent_id)
            return intent.status == 'succeeded', intent
        except stripe.error.StripeError:
            return False, None
    
    @transaction.atomic
    def post(self, request):
        serializer = CreateOrderSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        cart = self.get_cart(request)
        if not cart or cart.items.count() == 0:
            return Response(
                {'error': 'Cart is empty'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        for item in cart.items.all():
            if item.product.stock < item.quantity:
                return Response(
                    {'error': f'Insufficient stock for {item.product.name}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        payment_method = serializer.validated_data.get('payment_method', 'stripe')
        stripe_payment_intent_id = serializer.validated_data.get('stripe_payment_intent_id', '')
        
        if payment_method == 'cod':
            payment_status = 'pending'
            order_status = 'confirmed'
        else:
            # Stripe payment verification
            if stripe_payment_intent_id:
                is_valid, intent = self.verify_stripe_payment(stripe_payment_intent_id)
                if is_valid:
                    payment_status = 'paid'
                    order_status = 'confirmed'
                else:
                    return Response(
                        {'error': 'Payment verification failed'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            else:
                payment_status = 'pending'
                order_status = 'pending'
        
        order = Order.objects.create(
            user=request.user if request.user.is_authenticated else None,
            session_key=request.session.session_key if not request.user.is_authenticated else None,
            email=serializer.validated_data['email'],
            first_name=serializer.validated_data['first_name'],
            last_name=serializer.validated_data['last_name'],
            phone=serializer.validated_data.get('phone', ''),
            shipping_address=serializer.validated_data['shipping_address'],
            shipping_city=serializer.validated_data['shipping_city'],
            shipping_state=serializer.validated_data['shipping_state'],
            shipping_zip=serializer.validated_data['shipping_zip'],
            shipping_country=serializer.validated_data.get('shipping_country', 'India'),
            notes=serializer.validated_data.get('notes', ''),
            subtotal=cart.subtotal,
            total=cart.total,
            status=order_status,
            payment_status=payment_status,
            payment_method=payment_method,
            stripe_payment_intent_id=stripe_payment_intent_id,
        )
        
        for item in cart.items.all():
            OrderItem.objects.create(
                order=order,
                product=item.product,
                product_name=item.product.name,
                product_sku=item.product.sku,
                quantity=item.quantity,
                unit_price=item.unit_price,
                total_price=item.total_price
            )
            
            item.product.stock -= item.quantity
            item.product.save()
            
            track_interaction(
                request=request,
                product=item.product,
                interaction_type='purchase',
                metadata={
                    'quantity': item.quantity,
                    'order_id': str(order.id),
                    'order_number': order.order_number
                }
            )
        
        cart.items.all().delete()
        
        return Response(OrderSerializer(order).data, status=status.HTTP_201_CREATED)


class OrderListView(generics.ListAPIView):
    serializer_class = OrderListSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return Order.objects.filter(user=self.request.user)


class OrderDetailView(generics.RetrieveAPIView):
    serializer_class = OrderSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return Order.objects.filter(user=self.request.user)


class TrackOrderView(APIView):
    """Track order by order number (public endpoint)."""
    
    permission_classes = [permissions.AllowAny]
    
    def get(self, request, order_number):
        order = get_object_or_404(Order, order_number=order_number)
        
        return Response({
            'order_number': order.order_number,
            'status': order.status,
            'status_display': order.get_status_display(),
            'payment_status': order.payment_status,
            'payment_status_display': order.get_payment_status_display(),
            'payment_method': order.payment_method,
            'tracking_number': order.tracking_number,
            'estimated_delivery': order.estimated_delivery,
            'first_name': order.first_name,
            'last_name': order.last_name,
            'shipping_address': order.shipping_address,
            'shipping_city': order.shipping_city,
            'shipping_state': order.shipping_state,
            'shipping_zip': order.shipping_zip,
            'shipping_country': order.shipping_country,
            'total': str(order.total),
            'items': [
                {
                    'product_name': item.product_name,
                    'quantity': item.quantity,
                    'unit_price': str(item.unit_price),
                    'total_price': str(item.total_price),
                    'product_image': item.product.image.url if item.product and item.product.image else None
                }
                for item in order.items.all()
            ],
            'created_at': order.created_at,
            'updated_at': order.updated_at,
        })


# ================= Admin Views =================

class AdminOrderListView(generics.ListAPIView):
    queryset = Order.objects.all().prefetch_related('items', 'items__product')
    serializer_class = AdminOrderSerializer
    permission_classes = [permissions.IsAdminUser]
    filterset_fields = ['status', 'payment_status', 'payment_method']
    search_fields = ['order_number', 'email', 'first_name', 'last_name', 'phone']


class AdminOrderDetailView(generics.RetrieveUpdateAPIView):
    queryset = Order.objects.all().prefetch_related('items', 'items__product')
    serializer_class = AdminOrderSerializer
    permission_classes = [permissions.IsAdminUser]


class AdminUpdateOrderStatusView(APIView):
    permission_classes = [permissions.IsAdminUser]
    
    def patch(self, request, pk):
        order = get_object_or_404(Order, pk=pk)
        serializer = OrderStatusUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        if 'status' in serializer.validated_data:
            order.status = serializer.validated_data['status']
        
        if 'payment_status' in serializer.validated_data:
            order.payment_status = serializer.validated_data['payment_status']
        
        if 'tracking_number' in serializer.validated_data:
            order.tracking_number = serializer.validated_data['tracking_number']
        
        if 'estimated_delivery' in serializer.validated_data:
            order.estimated_delivery = serializer.validated_data['estimated_delivery']
        
        if 'admin_notes' in serializer.validated_data:
            order.admin_notes = serializer.validated_data['admin_notes']
        
        order.save()
        
        return Response(AdminOrderSerializer(order).data)


class AdminOrderStatsView(APIView):
    permission_classes = [permissions.IsAdminUser]
    
    def get(self, request):
        from django.db.models import Count, Sum
        from django.utils import timezone
        from datetime import timedelta
        
        days = int(request.query_params.get('days', 30))
        start_date = timezone.now() - timedelta(days=days)
        
        orders = Order.objects.filter(created_at__gte=start_date)
        
        total_orders = orders.count()
        total_revenue = orders.filter(payment_status='paid').aggregate(
            total=Sum('total'))['total'] or 0
        
        status_breakdown = orders.values('status').annotate(
            count=Count('id')
        ).order_by('status')
        
        payment_breakdown = orders.values('payment_status').annotate(
            count=Count('id')
        ).order_by('payment_status')
        
        recent_orders = OrderListSerializer(
            orders.order_by('-created_at')[:10], many=True
        ).data
        
        pending_confirmation = orders.filter(status='pending').count()
        pending_processing = orders.filter(status='confirmed').count()
        pending_shipping = orders.filter(status='processing').count()
        
        return Response({
            'period_days': days,
            'total_orders': total_orders,
            'total_revenue': float(total_revenue),
            'status_breakdown': list(status_breakdown),
            'payment_breakdown': list(payment_breakdown),
            'recent_orders': recent_orders,
            'pending_actions': {
                'needs_confirmation': pending_confirmation,
                'needs_processing': pending_processing,
                'needs_shipping': pending_shipping,
            }
        })


# ================= Export Views =================

class ExportOrdersExcelView(APIView):
    """Export all orders to Excel file with comprehensive details."""
    permission_classes = [permissions.IsAdminUser]
    
    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Get filters from query params
        status_filter = request.query_params.get('status')
        payment_filter = request.query_params.get('payment_status')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        # Build queryset
        orders = Order.objects.all().prefetch_related('items', 'items__product').order_by('-created_at')
        
        if status_filter:
            orders = orders.filter(status=status_filter)
        if payment_filter:
            orders = orders.filter(payment_status=payment_filter)
        if start_date:
            orders = orders.filter(created_at__date__gte=start_date)
        if end_date:
            orders = orders.filter(created_at__date__lte=end_date)
        
        # Create workbook
        wb = Workbook()
        
        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # === Orders Sheet ===
        ws_orders = wb.active
        ws_orders.title = "Orders"
        
        order_headers = [
            'Order Number', 'Date', 'Customer Name', 'Email', 'Phone',
            'Shipping Address', 'City', 'State', 'ZIP', 'Country',
            'Subtotal', 'Shipping', 'Tax', 'Total',
            'Status', 'Payment Status', 'Payment Method',
            'Stripe Payment ID', 'Tracking Number', 'Estimated Delivery', 'Admin Notes'
        ]
        
        for col, header in enumerate(order_headers, 1):
            cell = ws_orders.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        for row, order in enumerate(orders, 2):
            data = [
                order.order_number,
                order.created_at.strftime('%Y-%m-%d %H:%M'),
                f"{order.first_name} {order.last_name}",
                order.email,
                order.phone,
                order.shipping_address,
                order.shipping_city,
                order.shipping_state,
                order.shipping_zip,
                order.shipping_country,
                float(order.subtotal),
                float(order.shipping_cost),
                float(order.tax),
                float(order.total),
                order.get_status_display(),
                order.get_payment_status_display(),
                order.get_payment_method_display(),
                order.stripe_payment_intent_id or '',
                order.tracking_number or '',
                order.estimated_delivery.strftime('%Y-%m-%d') if order.estimated_delivery else '',
                order.admin_notes or ''
            ]
            for col, value in enumerate(data, 1):
                cell = ws_orders.cell(row=row, column=col, value=value)
                cell.border = thin_border
        
        for col in range(1, len(order_headers) + 1):
            ws_orders.column_dimensions[get_column_letter(col)].width = 15
        
        # === Order Items Sheet ===
        ws_items = wb.create_sheet(title="Order Items")
        
        item_headers = [
            'Order Number', 'Order Date', 'Customer', 'Product Name', 'SKU', 
            'Quantity', 'Unit Price', 'Total Price', 'Order Status', 'Payment Status'
        ]
        
        for col, header in enumerate(item_headers, 1):
            cell = ws_items.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row = 2
        for order in orders:
            for item in order.items.all():
                data = [
                    order.order_number,
                    order.created_at.strftime('%Y-%m-%d'),
                    f"{order.first_name} {order.last_name}",
                    item.product_name,
                    item.product_sku,
                    item.quantity,
                    float(item.unit_price),
                    float(item.total_price),
                    order.get_status_display(),
                    order.get_payment_status_display()
                ]
                for col, value in enumerate(data, 1):
                    cell = ws_items.cell(row=row, column=col, value=value)
                    cell.border = thin_border
                row += 1
        
        for col in range(1, len(item_headers) + 1):
            ws_items.column_dimensions[get_column_letter(col)].width = 18
        
        # === Summary Sheet ===
        ws_summary = wb.create_sheet(title="Summary")
        
        from django.db.models import Sum
        
        total_orders = orders.count()
        total_revenue = orders.filter(payment_status='paid').aggregate(total=Sum('total'))['total'] or 0
        total_items = sum(order.items.count() for order in orders)
        
        summary_data = [
            ('ORDERS EXPORT REPORT', ''),
            ('Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('', ''),
            ('OVERVIEW', ''),
            ('Total Orders', total_orders),
            ('Total Revenue', f"₹{total_revenue:,.2f}"),
            ('Total Items Sold', total_items),
            ('', ''),
            ('ORDER STATUS BREAKDOWN', ''),
            ('Pending', orders.filter(status='pending').count()),
            ('Confirmed', orders.filter(status='confirmed').count()),
            ('Processing', orders.filter(status='processing').count()),
            ('Shipped', orders.filter(status='shipped').count()),
            ('Out for Delivery', orders.filter(status='out_for_delivery').count()),
            ('Delivered', orders.filter(status='delivered').count()),
            ('Cancelled', orders.filter(status='cancelled').count()),
            ('Returned', orders.filter(status='returned').count()),
            ('', ''),
            ('PAYMENT STATUS BREAKDOWN', ''),
            ('Paid', orders.filter(payment_status='paid').count()),
            ('Pending Payment', orders.filter(payment_status='pending').count()),
            ('Failed', orders.filter(payment_status='failed').count()),
            ('Refunded', orders.filter(payment_status='refunded').count()),
            ('', ''),
            ('PAYMENT METHOD BREAKDOWN', ''),
            ('Stripe (Online)', orders.filter(payment_method='stripe').count()),
            ('Cash on Delivery', orders.filter(payment_method='cod').count()),
        ]
        
        for row, (label, value) in enumerate(summary_data, 1):
            cell_a = ws_summary.cell(row=row, column=1, value=label)
            cell_b = ws_summary.cell(row=row, column=2, value=value)
            if label in ['ORDERS EXPORT REPORT', 'OVERVIEW', 'ORDER STATUS BREAKDOWN', 'PAYMENT STATUS BREAKDOWN', 'PAYMENT METHOD BREAKDOWN']:
                cell_a.font = Font(bold=True, size=12)
        
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 25
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=orders_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        wb.save(response)
        return response


class ExportOrdersPDFView(APIView):
    """Export all orders to PDF file."""
    permission_classes = [permissions.IsAdminUser]
    
    def get(self, request):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER
        from django.db.models import Sum
        
        status_filter = request.query_params.get('status')
        payment_filter = request.query_params.get('payment_status')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        orders = Order.objects.all().prefetch_related('items', 'items__product').order_by('-created_at')
        
        if status_filter:
            orders = orders.filter(status=status_filter)
        if payment_filter:
            orders = orders.filter(payment_status=payment_filter)
        if start_date:
            orders = orders.filter(created_at__date__gte=start_date)
        if end_date:
            orders = orders.filter(created_at__date__lte=end_date)
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Heading1'], fontSize=24, spaceAfter=30,
            alignment=TA_CENTER, textColor=colors.HexColor('#1976D2')
        )
        
        section_style = ParagraphStyle(
            'SectionTitle', parent=styles['Heading2'], fontSize=16, spaceBefore=20,
            spaceAfter=10, textColor=colors.HexColor('#1976D2')
        )
        
        # Title
        elements.append(Paragraph("Orders Report", title_style))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Summary
        total_orders = orders.count()
        total_revenue = orders.filter(payment_status='paid').aggregate(total=Sum('total'))['total'] or 0
        
        elements.append(Paragraph("Summary", section_style))
        
        summary_data = [
            ['Total Orders', 'Total Revenue', 'Paid', 'Pending', 'Delivered', 'Cancelled'],
            [
                str(total_orders),
                f"₹{total_revenue:,.2f}",
                str(orders.filter(payment_status='paid').count()),
                str(orders.filter(payment_status='pending').count()),
                str(orders.filter(status='delivered').count()),
                str(orders.filter(status='cancelled').count())
            ]
        ]
        
        summary_table = Table(summary_data, colWidths=[1.6*inch]*6)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976D2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#E3F2FD')),
            ('GRID', (0, 0), (-1, -1), 1, colors.white),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # Orders Table
        elements.append(Paragraph("Order Details", section_style))
        
        order_headers = ['Order #', 'Date', 'Customer', 'Email', 'Total', 'Status', 'Payment', 'Tracking']
        order_data = [order_headers]
        
        for order in orders[:100]:  # Limit to 100 for PDF
            order_data.append([
                order.order_number,
                order.created_at.strftime('%Y-%m-%d'),
                f"{order.first_name} {order.last_name[:1]}.",
                order.email[:20] + '...' if len(order.email) > 20 else order.email,
                f"₹{order.total:,.0f}",
                order.get_status_display()[:10],
                order.get_payment_status_display()[:7],
                order.tracking_number[:12] if order.tracking_number else '-'
            ])
        
        order_table = Table(order_data, colWidths=[1.1*inch, 0.9*inch, 1.2*inch, 1.6*inch, 0.9*inch, 0.9*inch, 0.8*inch, 1.0*inch])
        order_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976D2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDBDBD')),
        ]))
        elements.append(order_table)
        
        if orders.count() > 100:
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(f"Showing 100 of {orders.count()} orders. Download Excel for complete data.", styles['Normal']))
        
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=orders_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        
        return response


class ExportAnalyticsExcelView(APIView):
    """Export comprehensive analytics to Excel with all details."""
    permission_classes = [permissions.IsAdminUser]
    
    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.db.models import Count, Sum
        from django.utils import timezone
        from datetime import timedelta
        from analytics.models import ProductInteraction
        from products.models import Product
        
        days = int(request.query_params.get('days', 30))
        start_date = timezone.now() - timedelta(days=days)
        
        wb = Workbook()
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # === Overview Sheet ===
        ws_overview = wb.active
        ws_overview.title = "Overview"
        
        orders = Order.objects.filter(created_at__gte=start_date)
        interactions = ProductInteraction.objects.filter(created_at__gte=start_date)
        
        total_revenue = orders.filter(payment_status='paid').aggregate(total=Sum('total'))['total'] or 0
        
        overview_data = [
            ('ANALYTICS EXPORT REPORT', ''),
            ('Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Period', f'Last {days} days'),
            ('', ''),
            ('KEY METRICS', ''),
            ('Total Orders', orders.count()),
            ('Total Revenue', f"₹{total_revenue:,.2f}"),
            ('Avg Order Value', f"₹{total_revenue / max(orders.count(), 1):,.2f}"),
            ('', ''),
            ('USER INTERACTIONS', ''),
            ('Product Views', interactions.filter(interaction_type='view').count()),
            ('Add to Cart', interactions.filter(interaction_type='add_to_cart').count()),
            ('Purchases', interactions.filter(interaction_type='purchase').count()),
            ('Cart to Purchase Rate', f"{(interactions.filter(interaction_type='purchase').count() / max(interactions.filter(interaction_type='add_to_cart').count(), 1)) * 100:.2f}%"),
            ('View to Purchase Rate', f"{(interactions.filter(interaction_type='purchase').count() / max(interactions.filter(interaction_type='view').count(), 1)) * 100:.2f}%"),
        ]
        
        for row, (label, value) in enumerate(overview_data, 1):
            cell_a = ws_overview.cell(row=row, column=1, value=label)
            cell_b = ws_overview.cell(row=row, column=2, value=value)
            if label in ['ANALYTICS EXPORT REPORT', 'KEY METRICS', 'USER INTERACTIONS']:
                cell_a.font = Font(bold=True, size=12)
        
        ws_overview.column_dimensions['A'].width = 25
        ws_overview.column_dimensions['B'].width = 30
        
        # === Products Performance Sheet ===
        ws_products = wb.create_sheet(title="Products Performance")
        
        product_headers = ['Product Name', 'SKU', 'Category', 'Price', 'Stock', 'Views', 'Cart Adds', 'Purchases', 'Revenue', 'Conversion Rate']
        for col, header in enumerate(product_headers, 1):
            cell = ws_products.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        products = Product.objects.all()
        row = 2
        for product in products:
            product_interactions = interactions.filter(product=product)
            views = product_interactions.filter(interaction_type='view').count()
            cart_adds = product_interactions.filter(interaction_type='add_to_cart').count()
            purchases = product_interactions.filter(interaction_type='purchase').count()
            product_revenue = OrderItem.objects.filter(
                product=product,
                order__created_at__gte=start_date,
                order__payment_status='paid'
            ).aggregate(total=Sum('total_price'))['total'] or 0
            
            conversion_rate = (purchases / max(views, 1)) * 100
            
            data = [
                product.name,
                product.sku,
                product.category.name if product.category else 'N/A',
                float(product.price),
                product.stock,
                views,
                cart_adds,
                purchases,
                float(product_revenue),
                f"{conversion_rate:.2f}%"
            ]
            for col, value in enumerate(data, 1):
                cell = ws_products.cell(row=row, column=col, value=value)
                cell.border = thin_border
            row += 1
        
        for col in range(1, len(product_headers) + 1):
            ws_products.column_dimensions[get_column_letter(col)].width = 15
        
        # === Daily Stats Sheet ===
        ws_daily = wb.create_sheet(title="Daily Stats")
        
        daily_headers = ['Date', 'Orders', 'Revenue', 'Items Sold', 'Views', 'Cart Adds', 'Purchases']
        for col, header in enumerate(daily_headers, 1):
            cell = ws_daily.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        from django.db.models.functions import TruncDate
        
        daily_orders = orders.annotate(date=TruncDate('created_at')).values('date').annotate(
            count=Count('id'),
            revenue=Sum('total')
        ).order_by('date')
        
        daily_views = interactions.filter(interaction_type='view').annotate(
            date=TruncDate('created_at')
        ).values('date').annotate(count=Count('id'))
        
        daily_cart = interactions.filter(interaction_type='add_to_cart').annotate(
            date=TruncDate('created_at')
        ).values('date').annotate(count=Count('id'))
        
        daily_purchases = interactions.filter(interaction_type='purchase').annotate(
            date=TruncDate('created_at')
        ).values('date').annotate(count=Count('id'))
        
        views_dict = {d['date']: d['count'] for d in daily_views}
        cart_dict = {d['date']: d['count'] for d in daily_cart}
        purchases_dict = {d['date']: d['count'] for d in daily_purchases}
        
        row = 2
        for day_data in daily_orders:
            date = day_data['date']
            items_sold = OrderItem.objects.filter(
                order__created_at__date=date,
                order__created_at__gte=start_date
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            data = [
                date.strftime('%Y-%m-%d') if date else '',
                day_data['count'],
                float(day_data['revenue'] or 0),
                items_sold,
                views_dict.get(date, 0),
                cart_dict.get(date, 0),
                purchases_dict.get(date, 0)
            ]
            for col, value in enumerate(data, 1):
                ws_daily.cell(row=row, column=col, value=value).border = thin_border
            row += 1
        
        for col in range(1, len(daily_headers) + 1):
            ws_daily.column_dimensions[get_column_letter(col)].width = 15
        
        # === Customers Sheet ===
        ws_customers = wb.create_sheet(title="Customers")
        
        customer_headers = ['Email', 'Name', 'Total Orders', 'Total Spent', 'Last Order', 'Avg Order Value']
        for col, header in enumerate(customer_headers, 1):
            cell = ws_customers.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        customer_data = orders.values('email', 'first_name', 'last_name').annotate(
            order_count=Count('id'),
            total_spent=Sum('total')
        ).order_by('-total_spent')
        
        row = 2
        for customer in customer_data:
            last_order = orders.filter(email=customer['email']).order_by('-created_at').first()
            avg_value = float(customer['total_spent'] or 0) / max(customer['order_count'], 1)
            
            data = [
                customer['email'],
                f"{customer['first_name']} {customer['last_name']}",
                customer['order_count'],
                float(customer['total_spent'] or 0),
                last_order.created_at.strftime('%Y-%m-%d') if last_order else '',
                avg_value
            ]
            for col, value in enumerate(data, 1):
                ws_customers.cell(row=row, column=col, value=value).border = thin_border
            row += 1
        
        for col in range(1, len(customer_headers) + 1):
            ws_customers.column_dimensions[get_column_letter(col)].width = 20
        
        # === All Interactions Sheet ===
        ws_interactions = wb.create_sheet(title="All Interactions")
        
        interaction_headers = ['Date', 'Time', 'Product', 'Interaction Type', 'User/Session', 'IP Address']
        for col, header in enumerate(interaction_headers, 1):
            cell = ws_interactions.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        row = 2
        for interaction in interactions.select_related('product', 'user').order_by('-created_at')[:1000]:
            data = [
                interaction.created_at.strftime('%Y-%m-%d'),
                interaction.created_at.strftime('%H:%M:%S'),
                interaction.product.name if interaction.product else 'N/A',
                interaction.interaction_type,
                interaction.user.email if interaction.user else interaction.session_key[:20] if interaction.session_key else 'Anonymous',
                interaction.ip_address or ''
            ]
            for col, value in enumerate(data, 1):
                ws_interactions.cell(row=row, column=col, value=value).border = thin_border
            row += 1
        
        for col in range(1, len(interaction_headers) + 1):
            ws_interactions.column_dimensions[get_column_letter(col)].width = 18
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=analytics_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        wb.save(response)
        return response


class ExportAnalyticsPDFView(APIView):
    """Export analytics summary to PDF."""
    permission_classes = [permissions.IsAdminUser]
    
    def get(self, request):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER
        from django.db.models import Count, Sum
        from django.utils import timezone
        from datetime import timedelta
        from analytics.models import ProductInteraction
        from products.models import Product
        
        days = int(request.query_params.get('days', 30))
        start_date = timezone.now() - timedelta(days=days)
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Heading1'], fontSize=24, spaceAfter=30,
            alignment=TA_CENTER, textColor=colors.HexColor('#1976D2')
        )
        
        section_style = ParagraphStyle(
            'SectionTitle', parent=styles['Heading2'], fontSize=16, spaceBefore=20,
            spaceAfter=10, textColor=colors.HexColor('#1976D2')
        )
        
        elements.append(Paragraph("Analytics Dashboard Report", title_style))
        elements.append(Paragraph(f"Period: Last {days} days | Generated: {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        orders = Order.objects.filter(created_at__gte=start_date)
        interactions = ProductInteraction.objects.filter(created_at__gte=start_date)
        
        total_revenue = orders.filter(payment_status='paid').aggregate(total=Sum('total'))['total'] or 0
        total_orders = orders.count()
        total_views = interactions.filter(interaction_type='view').count()
        total_cart = interactions.filter(interaction_type='add_to_cart').count()
        total_purchases = interactions.filter(interaction_type='purchase').count()
        
        # Key Metrics
        elements.append(Paragraph("Key Metrics", section_style))
        
        metrics_data = [
            ['Total Revenue', 'Total Orders', 'Avg Order', 'Views', 'Cart Adds', 'Conversion'],
            [
                f"₹{total_revenue:,.2f}",
                str(total_orders),
                f"₹{total_revenue / max(total_orders, 1):,.2f}",
                str(total_views),
                str(total_cart),
                f"{(total_purchases / max(total_views, 1)) * 100:.2f}%"
            ]
        ]
        
        metrics_table = Table(metrics_data, colWidths=[1.6*inch]*6)
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976D2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#E3F2FD')),
            ('GRID', (0, 0), (-1, -1), 1, colors.white),
        ]))
        elements.append(metrics_table)
        elements.append(Spacer(1, 20))
        
        # Top Products
        elements.append(Paragraph("Top 10 Products by Revenue", section_style))
        
        products = Product.objects.all()
        product_revenue = []
        for product in products:
            revenue = OrderItem.objects.filter(
                product=product,
                order__created_at__gte=start_date,
                order__payment_status='paid'
            ).aggregate(total=Sum('total_price'))['total'] or 0
            if revenue > 0:
                product_revenue.append((product, revenue))
        
        product_revenue.sort(key=lambda x: x[1], reverse=True)
        
        product_headers = ['Product', 'Revenue', 'Views', 'Purchases', 'Conversion']
        product_data = [product_headers]
        
        for product, revenue in product_revenue[:10]:
            product_interactions = interactions.filter(product=product)
            views = product_interactions.filter(interaction_type='view').count()
            purchases = product_interactions.filter(interaction_type='purchase').count()
            conversion = (purchases / max(views, 1)) * 100
            
            product_data.append([
                product.name[:25],
                f"₹{revenue:,.0f}",
                str(views),
                str(purchases),
                f"{conversion:.1f}%"
            ])
        
        if len(product_data) > 1:
            product_table = Table(product_data, colWidths=[3.5*inch, 1.2*inch, 1*inch, 1*inch, 1*inch])
            product_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDBDBD')),
            ]))
            elements.append(product_table)
        
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=analytics_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        
        return response


class ExportProductsExcelView(APIView):
    """Export all products to Excel."""
    permission_classes = [permissions.IsAdminUser]
    
    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from products.models import Product, Category
        
        wb = Workbook()
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Products Sheet
        ws_products = wb.active
        ws_products.title = "Products"
        
        headers = ['ID', 'SKU', 'Name', 'Category', 'Price', 'Discounted Price', 'Stock', 'Active', 'Featured', 'Created', 'Description']
        for col, header in enumerate(headers, 1):
            cell = ws_products.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        products = Product.objects.all().select_related('category')
        for row, product in enumerate(products, 2):
            data = [
                product.id,
                product.sku,
                product.name,
                product.category.name if product.category else 'N/A',
                float(product.price),
                float(product.discounted_price) if product.discounted_price else '',
                product.stock,
                'Yes' if product.is_active else 'No',
                'Yes' if product.is_featured else 'No',
                product.created_at.strftime('%Y-%m-%d'),
                product.description[:100] if product.description else ''
            ]
            for col, value in enumerate(data, 1):
                cell = ws_products.cell(row=row, column=col, value=value)
                cell.border = thin_border
        
        for col in range(1, len(headers) + 1):
            ws_products.column_dimensions[get_column_letter(col)].width = 15
        
        # Categories Sheet
        ws_categories = wb.create_sheet(title="Categories")
        
        cat_headers = ['ID', 'Name', 'Products Count', 'Active Products']
        for col, header in enumerate(cat_headers, 1):
            cell = ws_categories.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        categories = Category.objects.all()
        for row, category in enumerate(categories, 2):
            data = [
                category.id,
                category.name,
                category.products.count(),
                category.products.filter(is_active=True).count()
            ]
            for col, value in enumerate(data, 1):
                ws_categories.cell(row=row, column=col, value=value).border = thin_border
        
        for col in range(1, len(cat_headers) + 1):
            ws_categories.column_dimensions[get_column_letter(col)].width = 18
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=products_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        wb.save(response)
        return response
